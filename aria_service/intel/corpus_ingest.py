"""ARIA tiered corpus ingest.

Provides a cle

an interface for pushing proprietary or curated documents
into the RAG store with explicit provenance metadata (tier, source class,
region, etc). Used by the corpus-ingest CLI and the /api/aria/corpus/ingest
endpoint.

Why this exists separately from rag_store.ingest_document
══════════════════════════════════════════════════════════
The existing `read_document` flow ingests opportunistic documents (whatever
gets shared in WhatsApp / email) with minimal metadata. We need a parallel
path for *curated* corpus documents where the human ingesting them knows:

  - which tier the source belongs to (A primary / B secondary / C live / D proprietary)
  - what the source class is (SIPRI / RAND / DD report / template / etc)
  - what region or domain the document covers
  - whether it's CPLP-relevant

Storing tier metadata on every chunk is what lets retrieval prefer Tier A/D
sources over Tier B/C, and what lets the confidence footer cite the tier mix
of the supporting passages.

This module is INDEPENDENT of any deployed chat path. Adding new ingest
sources here cannot affect existing chat replies until the retrieval-side
tier-aware ranking is wired in (separate work).

Public API
══════════
    extract_text_from_bytes(raw_bytes, filename, mimetype) -> str
        Pure extraction. Handles PDF (PyMuPDF), DOCX (zipfile/xml fallback),
        XLSX (openpyxl), TXT/MD (decode). Raises ValueError if no extractor
        produces usable text.

    ingest_corpus_document(text, filename, tier, source_class, ...) -> dict
        Async wrapper around rag_store.ingest_document that injects the
        tier/provenance metadata into extra_metadata. Returns the same
        dict shape as ingest_document."""
from __future__ import annotations
from .engine_wiring import wire_success, wire_failure

import io
import logging
import re
import zipfile
from typing import Any

logger = logging.getLogger("aria.corpus_ingest")

# Tier vocabulary — MUST stay in sync with corpus_registry.py:VALID_TIERS.
# Past incident 2026-04-09: this set was missing B+, C+, and E even though
# corpus_registry.py defines all 7 tiers and the bulk-ingest CLIs accept
# them via --tier. The result was that Tier C+ ingest (the Lusophone moat,
# 19 sources) failed wholesale on the server side with HTTP 400 even
# though the registry and CLI both supported the tier name. The Tier B+
# (64 sources) and Tier E (15 sources) ingests would have failed the
# same way. Always update both sets together when adding tiers.
#
# 2026-04-09 corpus expansion v3: added A+ (real-time physical world
# tracking — most are deferred to Phase 4 because they need API keys
# or signal-stream infrastructure, but a few static sources go here)
# and F (private military companies / non-state armed actors).
VALID_TIERS = {"A", "A+", "B", "B+", "C", "C+", "D", "E", "F", "unknown"}

# ── R-F3376: the rights gate ───────────────────────────────────────────────
#
# Before this, ingest validated `tier` and nothing else. The docstring promised
# "full provenance metadata", but provenance meant WHO PUBLISHED it — never
# whether ARIA is allowed to HOLD or REPEAT it. ARIA's retrieved text reaches
# customer-facing output, so two classes must never be stored verbatim:
#   - third-party copyright (the vetting module already binds us to "clause
#     numbers ONLY — never store the standard's text (BSI copyright)"), and
#   - protectively marked material, which in defence/security work arrives
#     stamped OFFICIAL-SENSITIVE or "not for onward distribution".
#
# `rights` is REQUIRED: absent is not permissive. The two refused values have no
# override flag on purpose — `licensed` is the legitimate route for material we
# have paid for, so an absolute refusal cannot be argued around under deadline.
RIGHTS_VALUES = {
    "owned",                  # the operator owns it outright
    "public_domain",          # no rights subsist
    "open_licence",           # OGL / CC-BY / equivalent permitting storage + quotation
    "licensed",               # third-party, held under a licence — store, do not quote verbatim
    # Factual records ARIA composed from a third-party dataset's FIELDS (e.g.
    # sipri_ingest turns CSV columns into a sentence). Facts are not protected
    # and the expression is ours, so this is quotable — but it is labelled
    # distinctly rather than as `owned`, which would overstate our claim to the
    # underlying data.
    "derived_facts",
    "third_party_copyright",  # REFUSED
    "restricted",             # REFUSED — protective marking / confidential
}
REFUSED_RIGHTS = {"third_party_copyright", "restricted"}
# Verbatim quotation is safe only where rights permit reproduction. `licensed`
# material may be held and reasoned over but must be summarised, not quoted.
QUOTABLE_RIGHTS = {"owned", "public_domain", "open_licence", "derived_facts"}

# A declared label is a CLAIM. These patterns are EVIDENCE, and they override the
# claim: a document stamped "© BSI" or "OFFICIAL-SENSITIVE" is refused even when
# the uploader ticked "owned". Deliberately narrow — each pattern is a formal
# marking, not an ordinary English word, so normal prose about a "confidential
# investor briefing" does not trip it.
_MARKING_PATTERNS: tuple[tuple[str, str], ...] = (
    (r"©\s*BSI|\bBSI\s+copyright|\bBRITISH\s+STANDARD\b|\bBS\s?\d{4,5}(?:[-:]\d{4})?\b",
     "BSI / British Standard copyright notice"),
    (r"\bISO/IEC\s+\d+|\b©\s*ISO\b", "ISO copyright notice"),
    (r"\bOFFICIAL[-\s]SENSITIVE\b", "OFFICIAL-SENSITIVE marking"),
    (r"\b(TOP\s+SECRET|SECRET|NATO\s+RESTRICTED|NATO\s+CONFIDENTIAL)\b",
     "national security marking"),
    (r"\bnot\s+for\s+(onward\s+)?distribution\b", "distribution restriction"),
    (r"\bis\s+CONFIDENTIAL\b|\bSTRICTLY\s+CONFIDENTIAL\b", "confidentiality marking"),
    (r"\ball\s+rights\s+reserved\b", "all-rights-reserved notice"),
)


def detect_restricted_markings(text: str) -> list[str]:
    """Formal copyright / protective markings found in `text`.

    Evidence that contradicts a permissive declaration. Empty list means nothing
    was found — which is not proof the document is clear, only that the obvious
    stamps are absent.
    """
    if not isinstance(text, str) or not text.strip():
        return []
    import re as _re
    found: list[str] = []
    head = text[:20000]          # markings live on the cover/footer, not page 400
    for pattern, label in _MARKING_PATTERNS:
        m = _re.search(pattern, head, _re.I)
        if m:
            found.append(f"{label} ({m.group(0).strip()[:60]})")
    return found


def may_quote_verbatim(meta: Any) -> bool:
    """True only when the stored rights permit reproducing the text.

    Fail-closed: unknown, missing or malformed metadata is NOT quotable. Consumers
    that render citations should call this before emitting source text.
    """
    if not isinstance(meta, dict):
        return False
    return str(meta.get("rights") or "") in QUOTABLE_RIGHTS

# Sentinel for "we tried every extractor and got nothing useful"
class ExtractError(ValueError):
    """Raised when no extractor produces usable text from the input bytes."""


def extract_text_from_bytes(
    raw_bytes: bytes,
    filename: str,
    mimetype: str = "",
    *,
    max_chars: int = 200_000,
) -> str:
    """Extract plain text from a binary document.

    Pure function — no I/O beyond decoding the input bytes. Caller is
    responsible for reading the file and passing in the bytes.

    Falls through extractors in order: PDF → DOCX → XLSX → plain text decode.
    Raises ExtractError if none produce at least 30 chars.

    `max_chars` caps the output so a 500-page PDF doesn't blow up downstream
    chunking. The corpus ingest path uses 200k by default — much larger than
    the chat-path /api/aria/read-document cap of 15k, because corpus docs
    are *meant* to be long.
    """
    if not raw_bytes:
        raise ExtractError("empty input")
    fname = (filename or "").lower()
    mime = (mimetype or "").lower()

    extracted = ""

    # ── PDF ────────────────────────────────────────────────────────────
    if "pdf" in mime or fname.endswith(".pdf"):
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(stream=raw_bytes, filetype="pdf")
            extracted = "\n".join(page.get_text() for page in doc)
            doc.close()
        except ImportError:
            logger.warning("PyMuPDF not available — cannot extract PDF %s", filename)
        except Exception as e:
            logger.warning("PDF extraction failed for %s: %s", filename, e)

    # ── DOCX ───────────────────────────────────────────────────────────
    elif (
        "word" in mime
        or "officedocument.wordprocessingml" in mime
        or fname.endswith(".docx")
    ):
        try:
            zf = zipfile.ZipFile(io.BytesIO(raw_bytes))
            if "word/document.xml" in zf.namelist():
                xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
                extracted = re.sub(r"<[^>]+>", " ", xml)
                extracted = " ".join(extracted.split())
            zf.close()
        except Exception as e:
            logger.warning("DOCX extraction failed for %s: %s", filename, e)

    # ── XLSX ───────────────────────────────────────────────────────────
    elif (
        "spreadsheet" in mime
        or "officedocument.spreadsheetml" in mime
        or fname.endswith((".xlsx", ".xlsm"))
    ):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
            rows: list[str] = []
            for ws in wb.worksheets[:8]:  # cap at 8 sheets per workbook
                rows.append(f"--- Sheet: {ws.title} ---")
                for row in ws.iter_rows(max_row=2000, values_only=True):
                    rows.append(", ".join(str(c or "") for c in row))
            wb.close()
            extracted = "\n".join(rows)
        except Exception as e:
            logger.warning("XLSX extraction failed for %s: %s", filename, e)

    # ── Plain text / markdown ──────────────────────────────────────────
    elif fname.endswith((".txt", ".md", ".markdown", ".text")):
        try:
            extracted = raw_bytes.decode("utf-8", errors="replace")
        except Exception as e:
            logger.warning("Plain text decode failed for %s: %s", filename, e)

    # ── Last-resort: try utf-8 decode in case the caller mislabelled it ─
    if not extracted or len(extracted) < 30:
        try:
            maybe = raw_bytes.decode("utf-8", errors="ignore")
            if len(maybe.strip()) >= 30:
                extracted = maybe
        except Exception:
            pass

    if not extracted or len(extracted.strip()) < 30:
        raise ExtractError(f"no usable text extracted from {filename!r}")

    if len(extracted) > max_chars:
        logger.info(
            "corpus_ingest: truncating %s from %d to %d chars",
            filename, len(extracted), max_chars,
        )
        extracted = extracted[:max_chars]

    return extracted


async def ingest_corpus_document(
    text: str,
    *,
    filename: str,
    tier: str,
    source_class: str,
    region: str = "",
    cplp_relevant: bool = False,
    confidence: str = "",
    publication_date: str = "",
    notes: str = "",
    rights: str = "",
    rights_note: str = "",
    extra_metadata: dict | None = None,
) -> dict:
    """Push a corpus document into the RAG store with full provenance metadata.

    Parameters
    ----------
    text:
        The extracted plain text. Use extract_text_from_bytes() first.
    filename:
        Original filename — used as the chunk title and as part of the source id.
    tier:
        One of A/B/C/D/unknown. See module docstring for tier definitions.
    source_class:
        Free-form label for the publisher / origin (e.g. "SIPRI",
        "Arkmurus DD report", "Wassenaar control list").
    region:
        Region tag for filtering at retrieval time (e.g. "Africa",
        "Lusophone", "MENA").
    cplp_relevant:
        True if the document touches CPLP markets specifically. Used by the
        retrieval boost.
    confidence:
        Optional default confidence to attach if the source class warrants
        it (e.g. Tier A primary docs default to "CONFIRMED").
    publication_date:
        ISO date string if known.
    notes:
        Free-form note from the human ingesting the document.
    extra_metadata:
        Additional key/values merged into the chunk metadata last.

    Returns the dict shape from rag_store.ingest_document, with `tier`
    and `source_class` echoed for the caller's convenience.
    """
    if tier not in VALID_TIERS:
        raise ValueError(f"invalid tier {tier!r} — must be one of {sorted(VALID_TIERS)}")

    # ── R-F3376 rights gate — refuse BEFORE anything is written ────────────
    if not rights:
        raise ValueError(
            "rights is required — declare one of "
            f"{sorted(RIGHTS_VALUES)}. An unstated rights position is not a "
            "permissive one: ARIA's retrieved text reaches customer output."
        )
    if rights not in RIGHTS_VALUES:
        raise ValueError(f"invalid rights {rights!r} — must be one of {sorted(RIGHTS_VALUES)}")
    if rights in REFUSED_RIGHTS:
        raise ValueError(
            f"refusing to ingest {filename!r}: rights={rights!r}. Third-party "
            "copyright and restricted material must not be stored verbatim — the "
            "vetting module's rule is 'clause numbers ONLY, never the standard's "
            "text'. If a licence is held, declare rights='licensed' with a note."
        )
    if rights == "licensed" and not rights_note.strip():
        raise ValueError(
            "rights='licensed' requires rights_note saying under WHAT licence — "
            "an unnamed licence cannot be audited later."
        )
    # A declared label is a claim; a marking in the text is evidence. Evidence
    # wins, EXCEPT for 'licensed', where a copyright notice is exactly what a
    # licensed document is expected to carry.
    if rights != "licensed":
        markings = detect_restricted_markings(text)
        if markings:
            raise ValueError(
                f"refusing to ingest {filename!r}: declared rights={rights!r} but the "
                f"document carries {len(markings)} restrictive marking(s): "
                f"{'; '.join(markings[:3])}. If this is held under licence, declare "
                f"rights='licensed' with a note; otherwise it must not be stored."
            )

    from . import rag_store

    meta: dict[str, Any] = {
        "rights": rights,          # R-F3376 — carried to retrieval; see may_quote_verbatim
        "tier": tier,
        "source_class": source_class[:100],
        "region": region[:100],
        "cplp_relevant": bool(cplp_relevant),
    }
    if confidence:
        meta["confidence"] = confidence[:30]
    if publication_date:
        meta["publication_date"] = publication_date[:30]
    if notes:
        meta["notes"] = notes[:300]
    if rights_note:
        meta["rights_note"] = rights_note[:300]
    if extra_metadata:
        for k, v in extra_metadata.items():
            if isinstance(v, (str, int, float, bool)):
                meta[str(k)[:50]] = v if not isinstance(v, str) else v[:300]

    result = await rag_store.ingest_document(
        text=text,
        source=f"corpus:{tier}:{rights}:{source_class}:{filename}"[:300],
        source_type="corpus",
        title=filename,
        market=region,
        extra_metadata=meta,
    )
    result["tier"] = tier
    result["source_class"] = source_class
    result["rights"] = rights

    # R-F996 — wire to brain
    wire_success(
        module="corpus_ingest",
        summary="Corpus ingest",
        source_id="corpus_ingest:R-F996",
    )
    return result

# R-F2538: R-F2119 import-time wire_failure("module shutdown") block removed — it fired a FALSE engine_failure gap on every import (not at shutdown); do not re-add.
